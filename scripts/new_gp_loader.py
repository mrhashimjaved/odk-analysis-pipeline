"""
Utility for loading the New_GP re-randomisation reference and applying it to ODK rows.

New_GP codes
------------
11 = Group A - Continued (Responders)
12 = Group A - EASE Added (Non-Responders)
13 = Group A - EASE Control (Non-Responders)
21 = Group B - Continued (Responders)
22 = Group B - EASE Added (Non-Responders)
23 = Group B - EASE Control (Non-Responders)
"""

import os

DEFAULT_NEW_GP_FILE = "data/New_GP.xlsx"

NEW_GP_LABELS = {
    11: "11 - A Continued (Responders)",
    12: "12 - A EASE Added (Non-Responders)",
    13: "13 - A EASE Control (Non-Responders)",
    21: "21 - B Continued (Responders)",
    22: "22 - B EASE Added (Non-Responders)",
    23: "23 - B EASE Control (Non-Responders)",
}

# Full descriptive labels for use in reports
NEW_GP_FULL_LABELS = {
    11: "Group A - Continued (Responders)",
    12: "Group A - EASE Added (Non-Responders)",
    13: "Group A - EASE Control (Non-Responders)",
    21: "Group B - Continued (Responders)",
    22: "Group B - EASE Added (Non-Responders)",
    23: "Group B - EASE Control (Non-Responders)",
}

# Ordered list of all New_GP codes for consistent iteration
NEW_GP_ORDER = [11, 12, 13, 21, 22, 23]

# All 15 pairwise comparisons (group_1_code, group_2_code)
NEW_GP_PAIRS = [
    (a, b)
    for i, a in enumerate(NEW_GP_ORDER)
    for b in NEW_GP_ORDER[i + 1:]
]

# Candidate column names for child_id in ODK CSVs (tried in order)
CHILD_ID_CANDIDATES = [
    "child_id",
    "demo_a_-child_id",
    "demo_b_-child_id",
    "KEY",
    "_id",
    "meta-instanceID",
]


def find_child_id_col(column_names):
    """Return the first matching child_id column name, or None."""
    col_set = set(column_names)
    for candidate in CHILD_ID_CANDIDATES:
        if candidate in col_set:
            return candidate
    return None


def load_new_gp_reference(xlsx_path=DEFAULT_NEW_GP_FILE):
    """
    Load New_GP.xlsx and return a dict mapping child_id (str) → new_gp label (str).
    Returns an empty dict if the file is missing or unreadable.
    """
    if not os.path.exists(xlsx_path):
        print(f"[new_gp_loader] File not found: {xlsx_path} — New_GP grouping skipped.")
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if "child_id" not in headers or "New_GP" not in headers:
            print(f"[new_gp_loader] Required columns 'child_id' / 'New_GP' not found in {xlsx_path}.")
            return {}
        child_id_col = headers.index("child_id")
        new_gp_col = headers.index("New_GP")
        reference = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            child_id = row[child_id_col]
            new_gp = row[new_gp_col]
            if child_id is None or new_gp is None:
                continue
            try:
                gp_int = int(new_gp)
                label = NEW_GP_LABELS.get(gp_int, str(gp_int))
                reference[str(child_id).strip()] = label
            except (ValueError, TypeError):
                pass
        print(f"[new_gp_loader] Loaded {len(reference)} New_GP entries from {xlsx_path}.")
        return reference
    except Exception as exc:
        print(f"[new_gp_loader] Could not read {xlsx_path}: {exc} — New_GP grouping skipped.")
        return {}


def apply_new_gp(rows, reference, child_id_col):
    """
    Attach 'new_gp' label to each row in-place.
    Rows with no match get an empty string.
    """
    for row in rows:
        child_id = str(row.get(child_id_col, "") or "").strip()
        row["new_gp"] = reference.get(child_id, "")


def new_gp_label(code_int):
    """Return the short label for a numeric New_GP code."""
    return NEW_GP_LABELS.get(code_int, str(code_int))


def new_gp_full_label(code_int):
    """Return the full descriptive label for a numeric New_GP code."""
    return NEW_GP_FULL_LABELS.get(code_int, str(code_int))
